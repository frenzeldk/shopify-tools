"""Behavioural tests for the outbound, filename and workbook guards.

    python -m unittest discover -s web_tools/tests
"""
from __future__ import annotations

import os
import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import netguard  # noqa: E402
from purchase_order import security as po_security  # noqa: E402
from purchase_order.common import _cell_value, build_order_workbook  # noqa: E402
from purchase_order.common import OrderError  # noqa: E402
from purchase_order.templates import safe_attachment_name  # noqa: E402


class UrlValidation(unittest.TestCase):
    def test_rejects_non_https(self):
        with self.assertRaises(netguard.UnsafeURLError):
            netguard.validate_url("http://example.com/x")

    def test_rejects_embedded_credentials(self):
        with self.assertRaises(netguard.UnsafeURLError):
            netguard.validate_url("https://user:pass@example.com/x")

    def test_rejects_loopback(self):
        with self.assertRaises(netguard.UnsafeURLError):
            netguard.validate_url("https://127.0.0.1/admin")

    def test_rejects_private_ranges(self):
        for host in ("10.0.0.1", "192.168.1.1", "172.16.0.1", "100.64.0.1"):
            with self.subTest(host=host), self.assertRaises(netguard.UnsafeURLError):
                netguard.validate_url(f"https://{host}/")

    def test_rejects_cloud_metadata(self):
        with self.assertRaises(netguard.UnsafeURLError):
            netguard.validate_url("https://169.254.169.254/latest/meta-data/")

    def test_rejects_ipv6_loopback_and_mapped_v4(self):
        for host in ("[::1]", "[::ffff:127.0.0.1]", "[fd00::1]", "[fe80::1]"):
            with self.subTest(host=host), self.assertRaises(netguard.UnsafeURLError):
                netguard.validate_url(f"https://{host}/")

    def test_enforces_host_allowlist(self):
        with self.assertRaises(netguard.UnsafeURLError):
            netguard.validate_url("https://8.8.8.8/", allowed_hosts="vendor.example")

    def test_allows_listed_public_host(self):
        host, port, ips = netguard.validate_url(
            "https://8.8.8.8/v1", allowed_hosts="8.8.8.8"
        )
        self.assertEqual((host, port), ("8.8.8.8", 443))
        self.assertEqual(ips, ["8.8.8.8"])

    def test_allowlist_can_pin_a_port(self):
        with self.assertRaises(netguard.UnsafeURLError):
            netguard.validate_url("https://8.8.8.8:8443/", allowed_hosts="8.8.8.8:443")

    def test_timeout_is_clamped(self):
        self.assertEqual(
            netguard.clamp_timeout(10_000), netguard.MAX_TIMEOUT_SECONDS
        )


class OrderingPolicy(unittest.TestCase):
    def setUp(self):
        self._saved = {
            k: os.environ.get(k)
            for k in ("PO_ALLOWED_ENV_VARS", "PO_ALLOWED_API_HOSTS")
        }

    def tearDown(self):
        for key, value in self._saved.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value

    def test_env_lookup_denied_by_default(self):
        os.environ.pop("PO_ALLOWED_ENV_VARS", None)
        with self.assertRaises(po_security.PolicyError):
            po_security.resolve_env("FLASK_SECRET_KEY")

    def test_env_lookup_requires_explicit_allowlist(self):
        os.environ["PO_ALLOWED_ENV_VARS"] = "VENDOR_TOKEN"
        os.environ["VENDOR_TOKEN"] = "ok"
        self.assertEqual(po_security.resolve_env("VENDOR_TOKEN"), "ok")
        with self.assertRaises(po_security.PolicyError):
            po_security.resolve_env("OPENAI_API_KEY")

    def test_template_with_unlisted_env_is_rejected(self):
        os.environ["PO_ALLOWED_ENV_VARS"] = "VENDOR_TOKEN"
        os.environ["PO_ALLOWED_API_HOSTS"] = "vendor.example"
        template = {
            "type": "openapi",
            "base_url": "https://vendor.example",
            "place_order": {"path": "/orders"},
            "auth": {"type": "header", "headers": {"X-Leak": {"$env": "SHOPIFY_API_KEY"}}},
        }
        with self.assertRaises(po_security.PolicyError):
            po_security.validate_api_template(template)

    def test_template_with_unlisted_host_is_rejected(self):
        os.environ["PO_ALLOWED_ENV_VARS"] = "VENDOR_TOKEN"
        os.environ["PO_ALLOWED_API_HOSTS"] = "vendor.example"
        with self.assertRaises(po_security.PolicyError):
            po_security.validate_api_template(
                {"base_url": "https://attacker.example", "place_order": {"path": "/x"}}
            )

    def test_request_without_allowlist_is_refused(self):
        os.environ.pop("PO_ALLOWED_API_HOSTS", None)
        with self.assertRaises(po_security.PolicyError):
            po_security.safe_request("GET", "https://vendor.example/x")


class AttachmentNames(unittest.TestCase):
    def test_accepts_a_plain_name(self):
        self.assertEqual(
            safe_attachment_name("PO_{order_number}.xlsx"), "PO_{order_number}.xlsx"
        )

    def test_rejects_traversal_and_absolute_paths(self):
        for name in (
            "../../../etc/passwd",
            "/opt/shopify-tools/web_tools/purchase_orders.db",
            "..\\..\\windows\\system32\\x.xlsx",
            "C:\\temp\\x.xlsx",
            "sub/dir.xlsx",
            "bad\x00.xlsx",
            ".hidden.xlsx",
            "nul.xlsx",
        ):
            with self.subTest(name=name), self.assertRaises(OrderError):
                safe_attachment_name(name)


class WorkbookEscaping(unittest.TestCase):
    def test_formula_leading_characters_are_neutralised(self):
        for payload in ("=1+1", "+1", "-1", "@SUM(A1)", "  =cmd|' /c calc'!A0"):
            with self.subTest(payload=payload):
                self.assertTrue(str(_cell_value(payload)).startswith("'"))

    def test_ordinary_values_pass_through(self):
        self.assertEqual(_cell_value("TS-CTT-CO-01"), "TS-CTT-CO-01")
        self.assertEqual(_cell_value(3), 3)
        self.assertIsNone(_cell_value(None))

    def test_workbook_writes_text_not_formulas(self):
        from openpyxl import load_workbook
        import io

        data = build_order_workbook(
            [{"sku": "=HYPERLINK(\"http://x\")", "quantity": 2}],
            [{"field": "sku", "label": "SKU"}, {"field": "quantity", "label": "Qty"}],
        )
        sheet = load_workbook(io.BytesIO(data)).active
        self.assertEqual(sheet.cell(row=2, column=1).data_type, "s")
        self.assertTrue(sheet.cell(row=2, column=1).value.startswith("'"))


if __name__ == "__main__":
    unittest.main()
